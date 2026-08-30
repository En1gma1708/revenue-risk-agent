"""
Bulk version of the "try your own case" feature (custom_case.py) -- lets someone upload a
spreadsheet of cases (Excel .xlsx or plain .csv) instead of filling the single-case form
repeatedly. Built specifically so a demo can show "upload a file of scenarios, watch the agent
process each one live" -- the multi-case version of proving this isn't a replayed script.

Every row becomes a real CustomCaseInput, run through the exact same run_custom_case() path a
single form submission uses -- no separate code path to keep honest. Capped at MAX_ROWS per
upload: this runs each row synchronously against a real, already-scarce free-tier LLM quota, so
an accidental 500-row upload could burn a whole day's capacity in one request. The cap is a
guardrail on OUR OWN feature, consistent with the project's whole stance on bounded, deliberate
LLM usage rather than unbounded throughput.
"""

from __future__ import annotations

import csv
import io
from typing import Optional

import openpyxl
from pydantic import ValidationError

from custom_case import CustomCaseInput, run_custom_case
from models import Surface

MAX_ROWS = 15

# Column order for both the CSV and XLSX formats -- also what generate_sample_workbook() writes,
# so the sample file a visitor downloads is guaranteed to match what upload actually parses.
COLUMNS = [
    "surface", "customer_name", "amount_inr", "provider",
    "instrument_type", "error_reason", "attempt_number",
    "items", "abandonment_stage", "device", "minutes_since_abandon",
    "days_overdue", "contact_channel_pref",
]


def _row_to_input(row: dict[str, str]) -> CustomCaseInput:
    def _get(key: str) -> Optional[str]:
        v = row.get(key)
        if v is None:
            return None
        v = str(v).strip()
        return v or None

    kwargs: dict = {
        "surface": _get("surface"),
        "customer_name": _get("customer_name"),
        "amount_inr": float(_get("amount_inr") or 0),
    }
    if _get("provider"):
        kwargs["provider"] = _get("provider")
    if _get("instrument_type"):
        kwargs["instrument_type"] = _get("instrument_type")
    if _get("error_reason"):
        kwargs["error_reason"] = _get("error_reason")
    if _get("attempt_number"):
        kwargs["attempt_number"] = int(float(_get("attempt_number")))
    if _get("items"):
        kwargs["items"] = [s.strip() for s in _get("items").split(";") if s.strip()]
    if _get("abandonment_stage"):
        kwargs["abandonment_stage"] = _get("abandonment_stage")
    if _get("device"):
        kwargs["device"] = _get("device")
    if _get("minutes_since_abandon"):
        kwargs["minutes_since_abandon"] = float(_get("minutes_since_abandon"))
    if _get("days_overdue"):
        kwargs["days_overdue"] = int(float(_get("days_overdue")))
    if _get("contact_channel_pref"):
        kwargs["contact_channel_pref"] = _get("contact_channel_pref")

    return CustomCaseInput(**kwargs)


def parse_upload(filename: str, content: bytes) -> list[CustomCaseInput]:
    """Dispatches by extension. Raises ValueError with a row-specific message on bad data --
    caller (the endpoint) surfaces that directly rather than half-processing a broken file."""
    lower = filename.lower()
    rows: list[dict[str, str]] = []

    if lower.endswith(".csv"):
        text = content.decode("utf-8-sig")
        reader = csv.DictReader(io.StringIO(text))
        rows = list(reader)
    elif lower.endswith(".xlsx"):
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        header = [str(h).strip() if h is not None else "" for h in next(rows_iter)]
        for raw in rows_iter:
            if raw is None or all(c is None for c in raw):
                continue
            rows.append({header[i]: raw[i] for i in range(min(len(header), len(raw)))})
    else:
        raise ValueError(f"Unsupported file type: {filename} (expected .csv or .xlsx)")

    if not rows:
        raise ValueError("No data rows found in the uploaded file.")
    if len(rows) > MAX_ROWS:
        raise ValueError(f"Too many rows ({len(rows)}) -- this endpoint caps uploads at {MAX_ROWS} "
                          f"cases per request to protect shared free-tier LLM quota.")

    inputs: list[CustomCaseInput] = []
    for i, row in enumerate(rows, start=2):   # row 1 is the header
        try:
            inputs.append(_row_to_input(row))
        except (ValidationError, ValueError, TypeError) as e:
            raise ValueError(f"Row {i}: {e}")
    return inputs


def run_bulk(inputs: list[CustomCaseInput]) -> list[dict]:
    """Runs each input through the real agent loop, one at a time (sequential, not parallel --
    this is bounded LLM usage by design, not a throughput feature). A single row's failure
    (including a genuine quota-exhaustion HARD_STOP) doesn't stop the rest -- same per-case
    isolation principle as run_batch.py, at a much smaller scale."""
    results = []
    for i, payload in enumerate(inputs, start=1):
        try:
            case_id = run_custom_case(payload)
            results.append({"row": i, "customer_name": payload.customer_name, "case_id": case_id, "error": None})
        except Exception as e:  # noqa: BLE001 -- one row's unexpected failure must not kill the batch
            results.append({"row": i, "customer_name": payload.customer_name, "case_id": None, "error": str(e)})
    return results


# ---------------------------------------------------------------------------
# Sample template -- 9 realistic rows (3 per surface) so a demo can download this, open it,
# and re-upload it without inventing scenarios live. Amounts/attributes chosen to plausibly span
# different guardrail tiers using KNOWN thresholds (e.g. the real >Rs.50,000 approval rule) --
# not gaming the model's reasoning, just picking realistic scenarios likely to be interesting.
# ---------------------------------------------------------------------------

SAMPLE_ROWS = [
    ["payment_failure", "Priya Sharma", 3200, "", "card", "insufficient_funds", 1, "", "", "", "", "", ""],
    ["payment_failure", "Arjun Mehta", 68000, "", "upi", "invalid_vpa", 2, "", "", "", "", "", ""],
    ["payment_failure", "Kabir Singh", 1200, "", "card", "card_expired", 1, "", "", "", "", "", ""],
    ["checkout_abandonment", "Sanya Iyer", 9500, "", "", "", "", "", "otp_entry", "mobile_web", 20, "", ""],
    ["checkout_abandonment", "Rohan Das", 55000, "", "", "", "", "", "review", "desktop", 300, "", ""],
    ["checkout_abandonment", "Meera Nair", 2100, "", "", "", "", "", "instrument_select", "app", 10, "", ""],
    ["overdue_receivable", "Kavya Traders", 28000, "", "", "", "", "", "", "", "", 15, "email"],
    ["overdue_receivable", "Aarav Enterprises", 120000, "", "", "", "", "", "", "", "", 75, "call"],
    ["overdue_receivable", "Ishaan Textiles", 8000, "", "", "", "", "", "", "", "", 5, "whatsapp"],
]


def generate_sample_workbook(path: str) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "cases"
    ws.append(COLUMNS)
    for row in SAMPLE_ROWS:
        ws.append(row)
    for col_idx, col_name in enumerate(COLUMNS, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = max(14, len(col_name) + 2)
    wb.save(path)


if __name__ == "__main__":
    import sys
    out_path = sys.argv[1] if len(sys.argv) > 1 else "sample_cases.xlsx"
    generate_sample_workbook(out_path)
    print(f"Wrote {out_path} ({len(SAMPLE_ROWS)} sample rows)")
